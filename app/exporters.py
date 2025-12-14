from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

def export_xlsx_professional(path: Path, rows, headers, title="Relatório", meta=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")
    right = Alignment(horizontal="right")
    header_fill = PatternFill("solid", fgColor="1F2937")  # slate-ish
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=1, column=1).alignment = left

    # Meta lines
    r = 2
    if meta:
        for k, v in meta.items():
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            ws.cell(row=r, column=1, value=f"{k}: {v}").alignment = left
            r += 1
        r += 1  # blank line
    else:
        r = 3

    header_row = r
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data
    start_data_row = header_row + 1
    for i, row in enumerate(rows, start=start_data_row):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = border
            if c == 1:  # date
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = left
            elif c == 6:  # amount
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                cell.alignment = right
            else:
                cell.alignment = left

    # Auto width
    for col in range(1, len(headers) + 1):
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v))[:50] if isinstance(v, str) else len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)

    ws.freeze_panes = ws["A" + str(start_data_row)]
    wb.save(path)

def export_pdf_professional(path: Path, title: str, headers, rows, meta=None):
    doc = SimpleDocTemplate(str(path), pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    story.append(Spacer(1, 8))

    if meta:
        for k, v in meta.items():
            story.append(Paragraph(f"{k}: {v}", styles["Normal"]))
        story.append(Spacer(1, 10))

    data = [headers] + rows
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D1D5DB")),
        ("FONTSIZE", (0,1), (-1,-1), 9),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("ALIGN", (5,1), (5,-1), "RIGHT"),
    ]))
    story.append(tbl)

    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    doc.build(story)

def export_csv(path: Path, rows, headers):
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)